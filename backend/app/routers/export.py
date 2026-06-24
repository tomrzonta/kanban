"""Exportação de dados: planilha .xlsx pronta + datasets JSON para o Power BI.

- /excel              -> baixa um .xlsx formatado (Excel / Google Sheets).
- /dataset/*          -> JSON paginado, fonte VIVA para o Power BI
                         (Obter Dados -> Web). O botão "Atualizar" puxa dados frescos.
"""
import io
from datetime import datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy import text
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.core.database import get_db

router = APIRouter(prefix="/api/export", tags=["export"])

HEADER_FILL = PatternFill("solid", start_color="1D9E75")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")


def _style_header(ws, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 3, 50)


@router.get("/excel")
def export_excel(db: Session = Depends(get_db)):
    """Gera um .xlsx em memória com abas Tickets, Movimentações, Tempo e Resumo."""
    wb = Workbook()

    # Aba 1: Tickets (com prejuízo já calculado por linha).
    ws = wb.active
    ws.title = "Tickets"
    cols = ["ID", "Título", "Marca", "Modelo", "Distribuidora", "NF", "Origem",
            "Qtd", "Custo Unit.", "Prejuízo", "Coluna Atual", "Criado em"]
    ws.append(cols)
    rows = db.execute(text("""
        SELECT t.id, t.titulo, b.name AS marca, m.name AS modelo,
               t.distribuidora, t.numero_nf, t.origem, t.quantidade,
               t.custo_unitario, (t.custo_unitario * t.quantidade) AS prejuizo,
               c.name AS coluna, t.created_at
        FROM tickets t
        JOIN printer_models m ON m.id = t.printer_model_id
        JOIN printer_brands b ON b.id = m.brand_id
        JOIN columns c ON c.id = t.column_id
        ORDER BY t.created_at DESC
    """)).all()
    for r in rows:
        ws.append([str(r.id), r.titulo, r.marca, r.modelo, r.distribuidora,
                   r.numero_nf, r.origem, r.quantidade, float(r.custo_unitario),
                   float(r.prejuizo), r.coluna, r.created_at])
    _style_header(ws, len(cols)); _autosize(ws)

    # Aba 2: Movimentações (histórico bruto).
    ws2 = wb.create_sheet("Movimentacoes")
    ws2.append(["Ticket ID", "De", "Para", "Movido em"])
    for r in db.execute(text("""
        SELECT h.ticket_id, cf.name AS de, ct.name AS para, h.moved_at
        FROM ticket_history h
        LEFT JOIN columns cf ON cf.id = h.from_column_id
        JOIN columns ct ON ct.id = h.to_column_id
        ORDER BY h.ticket_id, h.moved_at
    """)).all():
        ws2.append([str(r.ticket_id), r.de or "(entrada)", r.para, r.moved_at])
    _style_header(ws2, 4); _autosize(ws2)

    # Aba 3: Tempo por etapa (calculado via LEAD).
    ws3 = wb.create_sheet("Tempo_Por_Etapa")
    ws3.append(["Ticket ID", "Coluna", "Aguardando Cliente?", "Horas na Etapa"])
    for r in db.execute(text("""
        WITH s AS (
            SELECT h.ticket_id, h.to_column_id,
                   EXTRACT(EPOCH FROM (
                       LEAD(h.moved_at) OVER (PARTITION BY h.ticket_id ORDER BY h.moved_at)
                       - h.moved_at)) / 3600 AS horas
            FROM ticket_history h)
        SELECT s.ticket_id, c.name, c.is_waiting_client, s.horas
        FROM s JOIN columns c ON c.id = s.to_column_id
        WHERE s.horas IS NOT NULL
    """)).all():
        ws3.append([str(r.ticket_id), r.name,
                    "Sim" if r.is_waiting_client else "Não", round(r.horas, 2)])
    _style_header(ws3, 4); _autosize(ws3)

    # Aba 4: Resumo com FÓRMULAS (recalcula ao editar as outras abas no Excel).
    ws4 = wb.create_sheet("Resumo")
    ws4.append(["Métrica", "Valor"])
    n = ws3.max_row
    ws4.append(["Total de tickets", "=COUNTA(Tickets!A2:A100000)"])
    ws4.append(["Prejuízo total (R$)", "=SUM(Tickets!J2:J100000)"])
    ws4.append(["Horas médias (análise interna)",
                f'=AVERAGEIF(Tempo_Por_Etapa!C2:C{n},"Não",Tempo_Por_Etapa!D2:D{n})'])
    ws4.append(["Horas médias (aguardando cliente)",
                f'=AVERAGEIF(Tempo_Por_Etapa!C2:C{n},"Sim",Tempo_Por_Etapa!D2:D{n})'])
    _style_header(ws4, 2); _autosize(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"tickets_garantia_{datetime.now():%Y%m%d_%H%M}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/dataset/tickets")
def dataset_tickets(skip: int = Query(0, ge=0), limit: int = Query(1000, le=5000),
                    db: Session = Depends(get_db)):
    """JSON paginado para o Power BI. skip/limit evitam payloads enormes / 429."""
    rows = db.execute(text("""
        SELECT t.id, t.titulo, b.name AS marca, m.name AS modelo,
               t.distribuidora, t.origem, t.quantidade, t.custo_unitario,
               (t.custo_unitario * t.quantidade) AS prejuizo,
               c.name AS coluna_atual, t.created_at
        FROM tickets t
        JOIN printer_models m ON m.id = t.printer_model_id
        JOIN printer_brands b ON b.id = m.brand_id
        JOIN columns c ON c.id = t.column_id
        ORDER BY t.created_at DESC LIMIT :l OFFSET :o
    """), {"l": limit, "o": skip}).mappings().all()
    total = db.execute(text("SELECT COUNT(*) FROM tickets")).scalar()
    return JSONResponse({"total": total, "skip": skip, "limit": limit,
                         "data": [dict(r) for r in rows]})


@router.get("/dataset/tempo-etapa")
def dataset_tempo_etapa(db: Session = Depends(get_db)):
    """Tempo por etapa já calculado — alimenta o gráfico de gargalos no Power BI."""
    rows = db.execute(text("""
        WITH s AS (
            SELECT h.ticket_id, h.to_column_id,
                   EXTRACT(EPOCH FROM (
                       LEAD(h.moved_at) OVER (PARTITION BY h.ticket_id ORDER BY h.moved_at)
                       - h.moved_at)) / 3600 AS horas
            FROM ticket_history h)
        SELECT s.ticket_id, c.name AS coluna, c.is_waiting_client,
               ROUND(s.horas::numeric, 2) AS horas
        FROM s JOIN columns c ON c.id = s.to_column_id
        WHERE s.horas IS NOT NULL
    """)).mappings().all()
    return JSONResponse({"data": [dict(r) for r in rows]})
